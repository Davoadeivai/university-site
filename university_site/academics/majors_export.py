"""فهرست رشته‌های پذیرش، برای دیدن و برداشتن.

داوطلبی که می‌خواهد رشته‌ها را با خانواده‌اش مرور کند یا کنار دستش
داشته باشد، نباید مجبور باشد صفحه را اسکرین‌شات بگیرد.

دو خروجی هست و هر دو عمداً روی سرور چیزی نصب نمی‌خواهند:

- اکسل، با openpyxl که از قبل برای خروجی پذیرش استفاده می‌شود.
- نسخهٔ چاپی، که همان HTML است با استایل چاپ؛ مرورگر خودش آن را
  PDF می‌کند. reportlab روی این هاست نیست و افزودنش برای یک جدول
  ساده، یک وابستگی تازه و یک ریسک تازه است.
"""
from __future__ import annotations

from django.http import HttpResponse
from django.shortcuts import render

from academics.models import Major

HEADERS = ['ردیف', 'دانشکده', 'گروه آموزشی', 'رشته', 'مقطع']


def _rows():
    """رشته‌های فعال، به ترتیب دانشکده و گروه."""
    majors = (
        Major.objects.filter(is_active=True)
        .select_related('department', 'group')
        .order_by('department__order', 'group__order', 'degree', 'name'))
    labels = dict(Major.DEGREE_CHOICES)
    for index, major in enumerate(majors, start=1):
        yield [
            index,
            major.department.name if major.department else '—',
            major.group.name if major.group else 'بدون گروه',
            major.name,
            labels.get(major.degree, major.degree),
        ]


def majors_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    book = Workbook()
    sheet = book.active
    sheet.title = 'رشته‌های پذیرش'
    sheet.sheet_view.rightToLeft = True      # وگرنه ستون‌ها وارونه می‌نشینند

    sheet.append(HEADERS)
    head_fill = PatternFill('solid', fgColor='7A1F2E')
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in _rows():
        sheet.append(row)

    for column, width in zip('ABCDE', (7, 30, 30, 42, 20)):
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = 'A2'                # سرستون هنگام اسکرول بماند

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument'
                     '.spreadsheetml.sheet')
    # نام فایل غیرلاتین باید filename* باشد، وگرنه بعضی مرورگرها
    # نام را خراب ذخیره می‌کنند
    response['Content-Disposition'] = (
        "attachment; filename=majors.xlsx; "
        "filename*=UTF-8''%D8%B1%D8%B4%D8%AA%D9%87-%D9%87%D8%A7.xlsx")
    book.save(response)
    return response


def majors_print(request):
    """نسخهٔ چاپی — مرورگر خودش PDF می‌کند."""
    return render(request, 'academics/majors_print.html', {
        'rows': list(_rows()),
        'page_title': 'رشته‌های پذیرش دانشجو',
    })
