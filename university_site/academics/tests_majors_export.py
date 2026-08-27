"""فهرست رشته‌های پذیرش: دیدن، چاپ، و برداشتن."""
import io

from django.core.cache import cache
from django.test import TestCase
from django.urls import reverse

from academics.models import AcademicGroup, Department, Major
from core.models import Slider


class MajorsExcelTests(TestCase):
    """داوطلب باید بتواند فهرست را کنار دستش داشته باشد."""

    @classmethod
    def setUpTestData(cls):
        faculty = Department.objects.create(
            name='دانشکده فنی و مهندسی', slug='fanni', order=1, is_active=True)
        group = AcademicGroup.objects.create(
            name='گروه کامپیوتر', slug='computer',
            department=faculty, order=1, is_active=True)
        Major.objects.create(
            name='مهندسی کامپیوتر', slug='m1', degree='bachelor_cont',
            department=faculty, group=group, is_active=True)
        Major.objects.create(
            name='کامپیوتر نرم‌افزار', slug='m2', degree='associate_cont',
            department=faculty, group=group, is_active=True)

    def setUp(self):
        cache.clear()

    def _book(self):
        from openpyxl import load_workbook

        response = self.client.get(reverse('academics:majors_excel'))
        return load_workbook(io.BytesIO(response.content))

    def test_it_downloads(self):
        response = self.client.get(reverse('academics:majors_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])

    def test_it_arrives_as_a_file_not_a_page(self):
        response = self.client.get(reverse('academics:majors_excel'))
        self.assertIn('attachment', response['Content-Disposition'])

    def test_the_persian_filename_survives(self):
        """بدون ‎filename*‎ بعضی مرورگرها نام را خراب ذخیره می‌کنند."""
        response = self.client.get(reverse('academics:majors_excel'))
        self.assertIn("filename*=UTF-8''", response['Content-Disposition'])

    def test_every_active_major_is_a_row(self):
        sheet = self._book().active
        self.assertEqual(sheet.max_row - 1, 2)

    def test_the_columns_are_named(self):
        sheet = self._book().active
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(
            headers, ['ردیف', 'دانشکده', 'گروه آموزشی', 'رشته', 'مقطع'])

    def test_the_sheet_reads_right_to_left(self):
        """وگرنه ستون‌ها وارونه می‌نشینند و جدول ناخوانا می‌شود."""
        self.assertTrue(self._book().active.sheet_view.rightToLeft)

    def test_the_header_stays_while_scrolling(self):
        self.assertEqual(self._book().active.freeze_panes, 'A2')

    def test_the_degree_is_spelled_out(self):
        """کد «bachelor_cont» به کار داوطلب نمی‌آید."""
        sheet = self._book().active
        values = [cell.value for row in sheet.iter_rows(min_row=2)
                  for cell in row]
        self.assertIn('کارشناسی پیوسته', values)
        self.assertNotIn('bachelor_cont', values)

    def test_an_inactive_major_is_left_out(self):
        Major.objects.filter(slug='m1').update(is_active=False)
        self.assertEqual(self._book().active.max_row - 1, 1)

    def test_a_major_without_a_group_still_appears(self):
        faculty = Department.objects.first()
        Major.objects.create(
            name='رشتهٔ بی‌گروه', slug='m3', degree='master',
            department=faculty, group=None, is_active=True)
        sheet = self._book().active
        values = [cell.value for row in sheet.iter_rows(min_row=2)
                  for cell in row]
        self.assertIn('رشتهٔ بی‌گروه', values)
        self.assertIn('بدون گروه', values)

    def test_an_empty_database_does_not_break_the_file(self):
        Major.objects.all().delete()
        self.assertEqual(
            self.client.get(reverse('academics:majors_excel')).status_code,
            200)


class MajorsPrintTests(TestCase):
    """نسخهٔ چاپی — مرورگر خودش PDF می‌کند."""

    @classmethod
    def setUpTestData(cls):
        faculty = Department.objects.create(
            name='دانشکده فنی', slug='fanni', order=1, is_active=True)
        group = AcademicGroup.objects.create(
            name='گروه کامپیوتر', slug='computer',
            department=faculty, order=1, is_active=True)
        Major.objects.create(
            name='مهندسی کامپیوتر', slug='m1', degree='bachelor_cont',
            department=faculty, group=group, is_active=True)

    def setUp(self):
        cache.clear()

    def _html(self):
        return self.client.get(
            reverse('academics:majors_print')).content.decode()

    def test_the_page_opens(self):
        self.assertEqual(
            self.client.get(reverse('academics:majors_print')).status_code,
            200)

    def test_the_table_carries_the_majors(self):
        html = self._html()
        self.assertIn('مهندسی کامپیوتر', html)
        self.assertIn('گروه کامپیوتر', html)
        self.assertIn('دانشکده فنی', html)

    def test_it_says_how_many(self):
        self.assertIn('1 رشته', self._html())

    def test_the_navigation_is_left_off_the_paper(self):
        """نوار بالا و فوتر روی کاغذ فقط جوهر هدر می‌دهند."""
        html = self._html()
        # لنگر روی خودِ قاعده، نه روی «اولین @media print»: بلوک
        # کوچکی برای اسکرول جدول زودتر در فایل می‌آید.
        rules = html.split('/* فقط جدول روی کاغذ می‌رود */')[1].split('}')[0]
        self.assertIn('display: none', rules)
        for part in ('header', 'footer', 'nav', '.print-actions'):
            self.assertIn(part, rules)

    def test_the_header_repeats_on_every_page(self):
        self.assertIn('display: table-header-group', self._html())

    def test_a_row_is_not_split_across_pages(self):
        self.assertIn('break-inside: avoid', self._html())

    def test_it_offers_both_ways_to_take_it(self):
        html = self._html()
        self.assertIn('window.print()', html)
        self.assertIn(reverse('academics:majors_excel'), html)

    def test_a_wide_table_scrolls_inside_itself(self):
        """جدول پهن نباید کل صفحه را افقی بکشد."""
        self.assertIn('overflow-x: auto', self._html())


class MajorsPageTakeAwayTests(TestCase):
    """راه رسیدن به این دو خروجی، از خود فهرست رشته‌هاست."""

    def setUp(self):
        cache.clear()

    def _html(self):
        return self.client.get(reverse('academics:majors')).content.decode()

    def test_both_buttons_are_offered(self):
        html = self._html()
        self.assertIn(reverse('academics:majors_print'), html)
        self.assertIn(reverse('academics:majors_excel'), html)

    def test_the_degree_filter_still_works(self):
        html = self._html()
        self.assertIn('tag-badge', html)

    def test_the_export_url_is_not_swallowed_by_the_detail_route(self):
        """«رشته‌ها/<slug>» زودتر بیاید، خروجی را یک رشته می‌بیند."""
        response = self.client.get(reverse('academics:majors_excel'))
        self.assertNotEqual(response.status_code, 404)
        self.assertIn('spreadsheetml', response['Content-Type'])


class SlideCallToActionTests(TestCase):
    """تنها نوشتهٔ روی اسلاید، و فقط وقتی مدیر پرش کند."""

    def setUp(self):
        cache.clear()
        for index in range(3):
            Slider.objects.create(
                title='', order=index, is_active=True,
                image='sliders/s%d.jpg' % index)

    def _hero(self):
        html = self.client.get(reverse('core:home')).content.decode()
        return html.split('id="heroTrack"')[1].split('/track')[0]

    def _run(self, *args):
        from io import StringIO

        from django.core.management import call_command

        out = StringIO()
        call_command('set_slide_cta', *args, stdout=out)
        return out.getvalue()

    def test_slides_stay_wordless_until_it_is_set(self):
        """موسسه خواسته بود اسلایدها بی‌نوشته باشند."""
        self.assertNotIn('slide-cta', self._hero())

    def test_the_command_puts_it_on_one_slide(self):
        self._run()
        cache.clear()
        self.assertEqual(self._hero().count('slide-cta'), 1)

    def test_it_carries_the_wording_the_institute_asked_for(self):
        self._run()
        cache.clear()
        self.assertIn('رشته‌های پذیرش دانشجو', self._hero())

    def test_it_leads_to_the_list_of_majors(self):
        self._run()
        cache.clear()
        self.assertIn(reverse('academics:majors'), self._hero())

    def test_it_lands_on_the_slide_you_name(self):
        self._run('--slide', '2')
        cache.clear()
        slides = self._hero().split('uni-hero-slide')
        self.assertIn('slide-cta', slides[2])
        self.assertNotIn('slide-cta', slides[1])

    def test_running_twice_does_not_leave_two_buttons(self):
        self._run()
        self._run('--slide', '3')
        cache.clear()
        self.assertEqual(self._hero().count('slide-cta'), 1)

    def test_clear_takes_it_off_again(self):
        self._run()
        self._run('--clear')
        cache.clear()
        self.assertNotIn('slide-cta', self._hero())

    def test_a_slide_number_past_the_end_uses_the_last_one(self):
        self._run('--slide', '99')
        cache.clear()
        self.assertEqual(self._hero().count('slide-cta'), 1)

    def test_it_says_so_when_there_are_no_slides(self):
        Slider.objects.all().delete()
        self.assertIn('اسلایدی ثبت نشده', self._run())

    def test_the_button_sits_on_the_left(self):
        """موسسه خواست سمت چپ باشد؛ در RTL آن ‎inline-end‎ است."""
        self._run()
        cache.clear()
        html = self.client.get(reverse('core:home')).content.decode()
        block = html.split('.slide-cta {')[1].split('}')[0]
        self.assertIn('inset-inline-end', block)
        self.assertNotIn('margin-inline: auto', block)

    def test_it_is_big_enough_to_be_seen_on_a_photograph(self):
        self._run()
        cache.clear()
        html = self.client.get(reverse('core:home')).content.decode()
        block = html.split('.slide-cta {')[1].split('}')[0]
        self.assertIn('clamp(15px, 1.7vw, 19px)', block)

    def test_it_carries_the_institute_gradient_not_a_flat_colour(self):
        """رنگ تخت روی عکس گم می‌شد."""
        self._run()
        cache.clear()
        html = self.client.get(reverse('core:home')).content.decode()
        block = html.split('.slide-cta {')[1].split('}')[0]
        self.assertIn('linear-gradient', block)
        self.assertIn('#a67c1f', block)

    def test_keyboard_focus_is_visible_on_any_photograph(self):
        self._run()
        cache.clear()
        html = self.client.get(reverse('core:home')).content.decode()
        self.assertIn('.slide-cta:focus-visible', html)

    def test_the_wording_can_be_changed(self):
        self._run('--label', 'رشته‌های موسسه')
        cache.clear()
        self.assertIn('رشته‌های موسسه', self._hero())
